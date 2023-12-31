import mysql.connector
from openpyxl import Workbook
from datetime import datetime, timedelta
from openpyxl.utils.dataframe import dataframe_to_rows


def start_date(connection):
    # вводим дата начала
    s_date = input('Введите дату в формате yyyy-mm-dd: ')
    all_tran(s_date, connection)


def all_tran(s_date, connection):
    # крайняя дата(правый предел)
    while s_date < '2023-06-10':
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                -- SQL запрос сюда --
         """)
        # извлекаем данные
        result = cursor.fetchall()
        filename = s_date
        workbook = Workbook()
        sheet = workbook.active
        last_row = sheet.max_row
        # определяем последнюю дату цикла
        last_date = sheet.cell(row=last_row, column=11).value
        for row in dataframe_to_rows(result, index=False, header=False):
            sheet.append(row)
        # сохраняем файл
        workbook.save(f"{filename}-{last_date}.xlsx")
        print(f"Результаты запроса сохранены в файл: {filename}")
        s_date = last_date + timedelta(days=1)
    cursor.close()


def machina(user, password, host, port, database):
    try:
        # Устанавливаем соединение с базой данных
        connection = mysql.connector.connect(
            user=user,
            password=password,
            host=host,
            port=port,
            database=database
        )
        connection.autocommit = True
        # Проверяем, что соединение установлено успешно
        if connection.is_connected():
            print("Соединение установлено")
            start_date(connection)
    except mysql.connector.Error as error:
        print("Ошибка при работе с MySQL:", error)
    finally:
        # Закрываем соединение с базой данных
        if connection and connection.is_connected():
            connection.close()
            print("Соединение с базой данных закрыто")

def config():
    print('Давайте подключимся к MySQL')
    user = input("Введите имя пользователя: ")
    password = input("Пароль: ")
    host = input("Хост: ")
    port = input("Порт: ")
    database = input("Название базы данных: ")
    machina(user, password, host, port, database)

if __name__ == '__main__':
    config()
